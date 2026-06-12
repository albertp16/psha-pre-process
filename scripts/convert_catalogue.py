"""Convert the PHIVOLCS "Earthquake Catalogue of the Philippines" workbook to JSON + GeoJSON.

Every non-empty cell in the workbook ends up in exactly one bucket so the
companion audit (scripts/audit_catalogue.py) can prove nothing was dropped:
  - metadata.sheets[*].preamble_cells  (title/region/period text above the header)
  - metadata.sheets[*].header_cells    (header row + unit/notes rows)
  - an event record field              (mapped data columns)
  - metadata.sheets[*].skipped_rows / stray_cells  (unexpected content -> audit failure)

Outputs:
  data/catalog.json     {metadata, events}  - full fidelity, includes intensity text
  data/catalog.geojson  FeatureCollection   - Mapbox-ready, no intensity text
"""
import argparse
import datetime as _dt
import hashlib
import json
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = REPO_ROOT / "data" / "source" / "Earthquake Catalogue of the Philippines.xlsx"

CANONICAL_HEADERS = {
    "year": "year",
    "month": "month",
    "day": "day",
    "hour": "hour",
    "minute": "minute",
    "second": "second",
    "north": "latitude",
    "east": "longitude",
    "depth": "depth_km",
    "ml": "ml",
    "mb": "mb",
    "ms": "ms",
    "mw": "mw",
    "intensity reports": "intensity_reports",
}

# Preferred magnitude per event: first available wins.
MAG_PRECEDENCE = [("mw", "Mw"), ("ms", "Ms"), ("mb", "Mb"), ("ml", "Ml")]

# Analysis floor. The workbook preamble states "Magnitude: 5.0 and above";
# events whose PREFERRED magnitude falls below this are excluded from
# `events` and the GeoJSON but kept verbatim in metadata.excluded_events so
# the capture audit still reconciles every workbook cell (nothing dropped).
MIN_MAGNITUDE = 5.0

INT_FIELDS = ("year", "month", "day", "hour", "minute")
COORD_FIELDS = ("latitude", "longitude", "depth_km")

FIELD_DESCRIPTIONS = {
    "id": "Sequential event id (sheet order, then row order)",
    "source_sheet": "Worksheet the event came from",
    "source_row": "1-based xlsx row number of the event",
    "year/month/day/hour/minute/second": "Origin time components as published (GMT)",
    "datetime_utc": "ISO 8601 origin time built from the components; null when components are invalid",
    "latitude": "Epicenter latitude, degrees N",
    "longitude": "Epicenter longitude, degrees E",
    "depth_km": "Focal depth, km",
    "ml/mb/ms/mw": "Published magnitudes by type (null when not reported); strings coerced to float",
    "mag": "Preferred magnitude: first available of Mw > Ms > Mb > Ml",
    "mag_type": "Which magnitude type populated `mag`",
    "intensity_reports": "Verbatim intensity report text (RF / PEIS scales)",
    "qa_flags": "Data-quality flags, e.g. invalid_datetime, ml_coerced_from_string",
}

NOTES = [
    "Workbook preamble states region 11.961-15.559N / 119.208-122.912E, but the data "
    "actually spans ~2-22N and ~116.3-133E (nationwide and beyond); the preamble appears "
    "to be a leftover from a website query and is preserved verbatim in preamble_cells.",
    "Sheet '1907-2018' preamble says 'Period Covered: 1910 to 2018' but contains events "
    "from 1907.",
    "Preamble states 'Magnitude: 5.0 and above'; some events carry individual magnitude "
    "types below 5.0 (e.g. Ml down to 3.0) alongside a larger magnitude of another type.",
    "Events whose PREFERRED magnitude is below 5.0 are excluded from `events` and the "
    "GeoJSON per that preamble floor; they are kept verbatim in metadata.excluded_events "
    "(count in metadata.n_excluded_below_min_mag) and remain capture-audited. Event ids "
    "keep their sequential slot, so an excluded event leaves a documented id gap.",
    "Origin times are GMT per the column header.",
]


def sha256_of(path):
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def cell_triples(row_vals, row_no):
    """All non-empty cells of a row as [row, column_letter, value]."""
    return [[row_no, get_column_letter(i + 1), v]
            for i, v in enumerate(row_vals) if v is not None]


def find_header(rows):
    """Return (header_index, {tuple_index: canonical_field}) for the row holding Year/Month."""
    for i, vals in enumerate(rows[:30]):
        strs = [str(v).strip() for v in vals if v is not None]
        if "Year" in strs and "Month" in strs:
            colmap = {}
            for idx, v in enumerate(vals):
                if v is None:
                    continue
                key = CANONICAL_HEADERS.get(str(v).strip().lower())
                if key is None:
                    raise ValueError(f"Unrecognized header {v!r} (column {get_column_letter(idx + 1)})")
                colmap[idx] = key
            return i, colmap
    raise ValueError("No header row containing 'Year' and 'Month' found")


def is_numberlike(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def build_datetime(ev, flags):
    """ISO 8601 string from the time components, or None with a qa flag."""
    try:
        y, mo, d = int(ev["year"]), int(ev["month"]), int(ev["day"])
        h = int(ev["hour"]) if ev["hour"] is not None else 0
        mi = int(ev["minute"]) if ev["minute"] is not None else 0
        s = float(ev["second"]) if ev["second"] is not None else 0.0
        sec = int(s)
        micro = int(round((s - sec) * 1e6))
        return _dt.datetime(y, mo, d, h, mi, sec, micro).isoformat() + "Z"
    except (TypeError, ValueError):
        flags.append("invalid_datetime")
        return None


def build_event(vals, colmap, sheet_name, row_no, eid):
    flags = []
    raw = {field: (vals[idx] if idx < len(vals) else None) for idx, field in colmap.items()}

    ev = {"id": eid, "source_sheet": sheet_name, "source_row": row_no}

    for f in INT_FIELDS:
        v = raw.get(f)
        ev[f] = int(v) if is_numberlike(v) else None
        if v is not None and not is_numberlike(v):
            flags.append(f"{f}_not_numeric")
    v = raw.get("second")
    ev["second"] = float(v) if is_numberlike(v) else None
    if raw.get("second") is not None and not is_numberlike(raw.get("second")):
        flags.append("second_not_numeric")

    ev["datetime_utc"] = build_datetime(ev, flags)

    for f in COORD_FIELDS:
        v = raw.get(f)
        if is_numberlike(v):
            ev[f] = float(v)
        else:
            ev[f] = None
            flags.append(f"missing_{f}" if v is None else f"{f}_not_numeric")

    for f in ("ml", "mb", "ms", "mw"):
        v = raw.get(f)
        if v is None:
            ev[f] = None
        elif is_numberlike(v):
            ev[f] = float(v)
        else:
            try:
                ev[f] = float(str(v).strip())
                flags.append(f"{f}_coerced_from_string")
            except ValueError:
                ev[f] = None
                ev.setdefault("unparsed", {})[f] = str(v)
                flags.append(f"{f}_unparseable")

    ev["mag"], ev["mag_type"] = None, None
    for f, label in MAG_PRECEDENCE:
        if ev.get(f) is not None:
            ev["mag"], ev["mag_type"] = ev[f], label
            break
    if ev["mag"] is None:
        flags.append("no_magnitude")

    ir = raw.get("intensity_reports")
    ev["intensity_reports"] = ir if isinstance(ir, str) else (str(ir) if ir is not None else None)

    ev["qa_flags"] = flags
    return ev


def convert_sheet(ws, start_id):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header_idx, colmap = find_header(rows)
    year_idx = next(idx for idx, f in colmap.items() if f == "year")

    preamble_cells = []
    for i in range(header_idx):
        preamble_cells.extend(cell_triples(rows[i], i + 1))

    # Header band: header row plus following unit/notes rows until the first numeric-Year row.
    data_start = None
    for i in range(header_idx + 1, len(rows)):
        v = rows[i][year_idx] if year_idx < len(rows[i]) else None
        if is_numberlike(v):
            data_start = i
            break
    if data_start is None:
        raise ValueError(f"Sheet {ws.title!r}: no data rows found")

    header_cells = []
    for i in range(header_idx, data_start):
        header_cells.extend(cell_triples(rows[i], i + 1))

    events, skipped_rows, stray_cells = [], [], []
    n_blank = 0
    eid = start_id
    for i in range(data_start, len(rows)):
        vals = rows[i]
        row_no = i + 1
        if all(v is None for v in vals):
            n_blank += 1
            continue
        if not is_numberlike(vals[year_idx] if year_idx < len(vals) else None):
            skipped_rows.append({"row": row_no, "cells": cell_triples(vals, row_no)})
            continue
        for idx, v in enumerate(vals):
            if v is not None and idx not in colmap:
                stray_cells.append([row_no, get_column_letter(idx + 1), v])
        events.append(build_event(vals, colmap, ws.title, row_no, eid))
        eid += 1

    ordered_cols = [colmap[idx] for idx in sorted(colmap)]
    sheet_meta = {
        "name": ws.title,
        "header_row": header_idx + 1,
        "data_start_row": data_start + 1,
        "data_end_row": len(rows),
        "max_column": ws.max_column,
        "columns": ordered_cols,
        "n_events": len(events),
        "n_blank_rows": n_blank,
        "preamble_cells": preamble_cells,
        "header_cells": header_cells,
        "skipped_rows": skipped_rows,
        "stray_cells": stray_cells,
    }
    return events, sheet_meta


def make_geojson(events):
    features = []
    for ev in events:
        props = {
            "id": ev["id"],
            "datetime_utc": ev["datetime_utc"],
            "year": ev["year"],
            "mag": ev["mag"],
            "mag_type": ev["mag_type"],
            "depth_km": ev["depth_km"],
        }
        for f in ("ml", "mb", "ms", "mw"):
            if ev[f] is not None:
                props[f] = ev[f]
        features.append({
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [ev["longitude"], ev["latitude"]]},
            "properties": props,
        })
    return {"type": "FeatureCollection", "features": features}


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--outdir", type=Path, default=REPO_ROOT / "data")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, read_only=True)
    all_events, sheets_meta = [], []
    for ws in wb.worksheets:
        events, meta = convert_sheet(ws, start_id=len(all_events) + 1)
        all_events.extend(events)
        sheets_meta.append(meta)
    wb.close()

    for ev in all_events:
        ev["id"] = f"PH-{ev['id']:04d}"

    # Apply the MIN_MAGNITUDE analysis floor after id assignment so the
    # surviving events keep their ids (the gap documents the exclusion).
    excluded_events = [ev for ev in all_events
                       if ev["mag"] is not None and ev["mag"] < MIN_MAGNITUDE]
    excluded_ids = {ev["id"] for ev in excluded_events}
    all_events = [ev for ev in all_events if ev["id"] not in excluded_ids]

    flag_counts = {}
    for ev in all_events:
        for fl in ev["qa_flags"]:
            flag_counts[fl] = flag_counts.get(fl, 0) + 1
    mag_type_counts = {}
    for ev in all_events:
        key = ev["mag_type"] or "none"
        mag_type_counts[key] = mag_type_counts.get(key, 0) + 1

    metadata = {
        "title": "Earthquake Catalogue of the Philippines (PHIVOLCS)",
        "source_agency": "PHIVOLCS",
        "source_file": args.xlsx.name,
        "source_sha256": sha256_of(args.xlsx),
        "generated_utc": _dt.datetime.now(_dt.timezone.utc).isoformat(),
        "time_basis": "GMT (per workbook header)",
        "magnitude_preference": [label for _, label in MAG_PRECEDENCE],
        "total_events": len(all_events),
        "min_magnitude": MIN_MAGNITUDE,
        "n_excluded_below_min_mag": len(excluded_events),
        "excluded_events": excluded_events,
        "mag_type_counts": mag_type_counts,
        "qa_flag_counts": flag_counts,
        "year_min": min(ev["year"] for ev in all_events if ev["year"] is not None),
        "year_max": max(ev["year"] for ev in all_events if ev["year"] is not None),
        "notes": NOTES,
        "field_descriptions": FIELD_DESCRIPTIONS,
        "sheets": sheets_meta,
    }

    args.outdir.mkdir(parents=True, exist_ok=True)
    catalog_path = args.outdir / "catalog.json"
    geojson_path = args.outdir / "catalog.geojson"
    with open(catalog_path, "w", encoding="utf-8") as fh:
        json.dump({"metadata": metadata, "events": all_events}, fh, ensure_ascii=False, indent=1)
    with open(geojson_path, "w", encoding="utf-8") as fh:
        json.dump(make_geojson(all_events), fh, ensure_ascii=False)

    print(f"source: {args.xlsx}")
    for m in sheets_meta:
        print(f"  sheet '{m['name']}': {m['n_events']} events "
              f"(rows {m['data_start_row']}-{m['data_end_row']}, header row {m['header_row']}), "
              f"{len(m['skipped_rows'])} skipped, {len(m['stray_cells'])} stray cells")
    print(f"total events: {len(all_events)} "
          f"(excluded below M{MIN_MAGNITUDE:.1f}: {len(excluded_events)})")
    print(f"mag_type counts: {mag_type_counts}")
    print(f"qa flags: {flag_counts}")
    print(f"wrote {catalog_path} ({catalog_path.stat().st_size:,} bytes)")
    print(f"wrote {geojson_path} ({geojson_path.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
