"""Audit that data/catalog.json + data/catalog.geojson capture ALL data in the PHIVOLCS workbook.

Independently re-reads the xlsx (full width, every cell) and verifies:
  1. Counts        - per-sheet data rows == events; geojson features == events
  2. Reconciliation- every mapped data cell equals the event field (float tol 1e-9, text exact)
  3. Partition     - every non-empty cell in the workbook is accounted for: preamble_cells,
                     header_cells, an event field, or (failure) skipped/stray
  4. GeoJSON       - coordinates/properties mirror the events; intensity text excluded by design
  5. Validity      - coordinate/magnitude/depth ranges, datetime validity, duplicates
  6. Ledger        - for the known source file (by sha256), the exact expected anomalies

Exit code 0 = pass, 1 = capture failure. Writes reports/audit_report.md + reports/audit.json.
"""
import datetime as _dt
import hashlib
import json
import random
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
XLSX = REPO_ROOT / "data" / "source" / "Earthquake Catalogue of the Philippines.xlsx"
CATALOG = REPO_ROOT / "data" / "catalog.json"
GEOJSON = REPO_ROOT / "data" / "catalog.geojson"
REPORTS = REPO_ROOT / "reports"

TOL = 1e-9

CANONICAL_HEADERS = {
    "year": "year", "month": "month", "day": "day", "hour": "hour",
    "minute": "minute", "second": "second",
    "north": "latitude", "east": "longitude", "depth": "depth_km",
    "ml": "ml", "mb": "mb", "ms": "ms", "mw": "mw",
    "intensity reports": "intensity_reports",
}
FLOAT_FIELDS = ("second", "latitude", "longitude", "depth_km", "ml", "mb", "ms", "mw")
INT_FIELDS = ("year", "month", "day", "hour", "minute")

# Exact anomaly ledger expected for the known source workbook (keyed by sha256).
# For any other source file these are reported but not asserted.
KNOWN_SHA = "e2a22971d6f98dfac02b6d081099f7e365b40dc9ca2e5555d3d58f893ba610a3"
EXPECTED_LEDGERS = {
    KNOWN_SHA: {
        "total_events": 3577,
        "per_sheet": {"1907-2018": 3032, "2019-2025": 545},
        "qa_flag_counts": {
            "ml_coerced_from_string": 1250,
            "mb_coerced_from_string": 1855,
            "invalid_datetime": 1,
        },
        "invalid_datetime_rows": [["2019-2025", 329]],
        "mag_type_counts": {"Ms": 1686, "Mb": 1035, "Ml": 446, "Mw": 410},
        "n_duplicate_pairs": 24,
    }
}

failures, warnings, info = [], [], []


def fail(msg):
    failures.append(msg)


def warn(msg):
    warnings.append(msg)


def is_numberlike(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def eq(cell, val, float_ok):
    """Compare an xlsx cell value against a JSON field value."""
    if cell is None and val is None:
        return True
    if cell is None or val is None:
        return False
    if float_ok:
        try:
            return abs(float(str(cell).strip() if isinstance(cell, str) else cell) - float(val)) <= TOL
        except (TypeError, ValueError):
            return False
    return str(cell) == str(val)


def sha256_of(path):
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def find_header(rows, sheet):
    for i, vals in enumerate(rows[:30]):
        strs = [str(v).strip() for v in vals if v is not None]
        if "Year" in strs and "Month" in strs:
            colmap = {}
            for idx, v in enumerate(vals):
                if v is None:
                    continue
                key = CANONICAL_HEADERS.get(str(v).strip().lower())
                if key is None:
                    fail(f"{sheet}: unrecognized header {v!r} at column {get_column_letter(idx + 1)}")
                    continue
                colmap[idx] = key
            return i, colmap
    fail(f"{sheet}: header row not found")
    return None, {}


def main():
    catalog = json.loads(CATALOG.read_text(encoding="utf-8"))
    geo = json.loads(GEOJSON.read_text(encoding="utf-8"))
    meta = catalog["metadata"]
    events = catalog["events"]

    # ── 0. Source identity ──
    actual_sha = sha256_of(XLSX)
    if meta.get("source_sha256") != actual_sha:
        fail(f"source sha256 mismatch: metadata={meta.get('source_sha256')} actual={actual_sha}")
    known = actual_sha == KNOWN_SHA and meta.get("source_sha256") == KNOWN_SHA

    ev_by_sheet_row = {(ev["source_sheet"], ev["source_row"]): ev for ev in events}
    if len(ev_by_sheet_row) != len(events):
        fail("duplicate (source_sheet, source_row) keys among events")

    sheet_meta = {m["name"]: m for m in meta["sheets"]}
    wb = openpyxl.load_workbook(XLSX, read_only=True)

    n_data_rows_total = 0
    n_cells_nonempty = 0
    n_cells_reconciled = 0
    recon_mismatches = []
    unclaimed = []
    claimed_missing = []
    ghost_max_col_seen = 0

    for ws in wb.worksheets:
        sheet = ws.title
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        header_idx, colmap = find_header(rows, sheet)
        if header_idx is None:
            continue
        year_idx = next(idx for idx, f in colmap.items() if f == "year")
        sm = sheet_meta.get(sheet)
        if sm is None:
            fail(f"{sheet}: sheet missing from metadata.sheets")
            continue

        # Claimed metadata cells: (row, col_letter) -> value
        claimed = {}
        for t in sm["preamble_cells"] + sm["header_cells"] + sm["stray_cells"]:
            claimed[(t[0], t[1])] = t[2]
        for sk in sm["skipped_rows"]:
            for t in sk["cells"]:
                claimed[(t[0], t[1])] = t[2]
        if sm["stray_cells"]:
            fail(f"{sheet}: converter recorded {len(sm['stray_cells'])} stray cells (data outside mapped columns)")
        if sm["skipped_rows"]:
            fail(f"{sheet}: converter skipped {len(sm['skipped_rows'])} non-empty rows")

        # Determine data start the same way the converter does (first numeric Year)
        data_start = None
        for i in range(header_idx + 1, len(rows)):
            if is_numberlike(rows[i][year_idx] if year_idx < len(rows[i]) else None):
                data_start = i
                break

        n_sheet_data_rows = 0
        seen_claimed = set()
        for i, vals in enumerate(rows):
            row_no = i + 1
            ghost_max_col_seen = max(ghost_max_col_seen, len(vals))
            is_data_row = (data_start is not None and i >= data_start
                           and is_numberlike(vals[year_idx] if year_idx < len(vals) else None))
            if is_data_row:
                n_sheet_data_rows += 1
                ev = ev_by_sheet_row.get((sheet, row_no))
                if ev is None:
                    fail(f"{sheet} row {row_no}: data row has NO corresponding event in catalog.json")
                    continue
                # Reconcile every mapped column (including empty->null)
                for idx, field in colmap.items():
                    cell = vals[idx] if idx < len(vals) else None
                    if cell is not None:
                        n_cells_nonempty += 1
                    val = ev.get(field)
                    ok = eq(cell, val, field in FLOAT_FIELDS or field in INT_FIELDS)
                    if not ok and field in ("ml", "mb", "ms", "mw") and ev.get("unparsed", {}).get(field):
                        ok = str(cell) == ev["unparsed"][field]
                    if field == "intensity_reports":
                        ok = (cell is None and val is None) or (cell is not None and str(cell) == val)
                    if ok:
                        n_cells_reconciled += 1
                    else:
                        recon_mismatches.append(f"{sheet}!{get_column_letter(idx + 1)}{row_no} "
                                                f"({field}): xlsx={cell!r} json={val!r}")
                # Stray content in unmapped columns of a data row
                for idx, v in enumerate(vals):
                    if v is not None and idx not in colmap:
                        key = (row_no, get_column_letter(idx + 1))
                        if key not in claimed:
                            unclaimed.append(f"{sheet}!{key[1]}{key[0]} = {v!r}")
                        else:
                            seen_claimed.add(key)
            else:
                # Non-data row: every non-empty cell must be claimed by metadata with equal value
                for idx, v in enumerate(vals):
                    if v is None:
                        continue
                    n_cells_nonempty += 1
                    key = (row_no, get_column_letter(idx + 1))
                    if key not in claimed:
                        unclaimed.append(f"{sheet}!{key[1]}{key[0]} = {v!r}")
                    else:
                        seen_claimed.add(key)
                        cv = claimed[key]
                        if not (eq(v, cv, isinstance(v, (int, float)))
                                or str(v) == str(cv)):
                            recon_mismatches.append(f"{sheet}!{key[1]}{key[0]} metadata cell: "
                                                    f"xlsx={v!r} json={cv!r}")
        # Reverse: metadata must not claim cells that don't exist
        for key, cv in claimed.items():
            if key not in seen_claimed:
                claimed_missing.append(f"{sheet}!{key[1]}{key[0]} = {cv!r} (claimed in metadata, absent in xlsx)")

        n_data_rows_total += n_sheet_data_rows
        if n_sheet_data_rows != sm["n_events"]:
            fail(f"{sheet}: xlsx data rows={n_sheet_data_rows} but metadata says n_events={sm['n_events']}")
        ev_in_sheet = sum(1 for ev in events if ev["source_sheet"] == sheet)
        if n_sheet_data_rows != ev_in_sheet:
            fail(f"{sheet}: xlsx data rows={n_sheet_data_rows} but events in catalog={ev_in_sheet}")
        info.append(f"{sheet}: {n_sheet_data_rows} data rows, all reconciled against events")

    wb.close()

    if unclaimed:
        fail(f"{len(unclaimed)} non-empty workbook cells NOT captured anywhere; first: {unclaimed[:10]}")
    if claimed_missing:
        fail(f"{len(claimed_missing)} metadata cells not present in workbook; first: {claimed_missing[:10]}")
    if recon_mismatches:
        fail(f"{len(recon_mismatches)} cell/field mismatches; first: {recon_mismatches[:10]}")
    if n_data_rows_total != len(events):
        fail(f"total xlsx data rows={n_data_rows_total} != events={len(events)}")
    if meta["total_events"] != len(events):
        fail(f"metadata.total_events={meta['total_events']} != len(events)={len(events)}")

    # ── 4. GeoJSON cross-check ──
    feats = geo["features"]
    if len(feats) != len(events):
        fail(f"geojson features={len(feats)} != events={len(events)}")
    for ftr, ev in zip(feats, events):
        p = ftr["properties"]
        c = ftr["geometry"]["coordinates"]
        if p.get("id") != ev["id"]:
            fail(f"geojson order mismatch at {p.get('id')} vs {ev['id']}")
            break
        if not (eq(c[0], ev["longitude"], True) and eq(c[1], ev["latitude"], True)):
            fail(f"{ev['id']}: geojson coords {c} != event lon/lat")
        for k in ("datetime_utc", "year", "mag", "mag_type", "depth_km"):
            if p.get(k) != ev[k]:
                fail(f"{ev['id']}: geojson prop {k}={p.get(k)!r} != event {ev[k]!r}")
        for k in ("ml", "mb", "ms", "mw"):
            if (k in p) != (ev[k] is not None) or (k in p and p[k] != ev[k]):
                fail(f"{ev['id']}: geojson prop {k} inconsistent with event")
        if "intensity_reports" in p:
            fail(f"{ev['id']}: intensity_reports must not be in geojson (by design)")

    # ── 5. Validity / sanity ──
    n_null_coord = sum(1 for ev in events if ev["latitude"] is None or ev["longitude"] is None)
    if n_null_coord:
        fail(f"{n_null_coord} events with null coordinates")
    lats = [ev["latitude"] for ev in events]
    lons = [ev["longitude"] for ev in events]
    depths = [ev["depth_km"] for ev in events if ev["depth_km"] is not None]
    mags = [ev["mag"] for ev in events if ev["mag"] is not None]
    if min(lats) < 0 or max(lats) > 25:
        warn(f"latitude outside [0,25]: {min(lats)}..{max(lats)}")
    if min(lons) < 110 or max(lons) > 135:
        warn(f"longitude outside [110,135]: {min(lons)}..{max(lons)}")
    if min(depths) < 0 or max(depths) > 700:
        warn(f"depth outside [0,700]: {min(depths)}..{max(depths)}")
    if min(mags) < 2.5 or max(mags) > 9:
        warn(f"preferred magnitude outside [2.5,9]: {min(mags)}..{max(mags)}")
    n_no_mag = sum(1 for ev in events if ev["mag"] is None)
    if n_no_mag:
        fail(f"{n_no_mag} events without any magnitude")
    bad_dt = [(ev["source_sheet"], ev["source_row"]) for ev in events if ev["datetime_utc"] is None]

    # Duplicates: identical origin time + location
    seen, dups = {}, []
    for ev in events:
        key = (ev["year"], ev["month"], ev["day"], ev["hour"], ev["minute"],
               ev["second"], ev["latitude"], ev["longitude"])
        if key in seen:
            dups.append((seen[key], ev["id"]))
        else:
            seen[key] = ev["id"]
    if dups:
        warn(f"{len(dups)} exact duplicate origin-time+location pairs: {dups[:5]}")

    sub_m5 = sum(1 for ev in events if ev["mag"] is not None and ev["mag"] < 5.0)
    info.append(f"events with preferred mag < 5.0: {sub_m5} (workbook preamble nominally M>=5.0)")

    # ── 6. Expected-anomaly ledger ──
    mag_type_counts = {}
    for ev in events:
        mag_type_counts[ev["mag_type"] or "none"] = mag_type_counts.get(ev["mag_type"] or "none", 0) + 1
    flag_counts = {}
    for ev in events:
        for fl in ev["qa_flags"]:
            flag_counts[fl] = flag_counts.get(fl, 0) + 1
    ledger = {
        "total_events": len(events),
        "per_sheet": {m["name"]: m["n_events"] for m in meta["sheets"]},
        "qa_flag_counts": flag_counts,
        "invalid_datetime_rows": [list(x) for x in bad_dt],
        "mag_type_counts": mag_type_counts,
        "n_duplicate_pairs": len(dups),
    }
    if known:
        exp = EXPECTED_LEDGERS[KNOWN_SHA]
        for k, v in exp.items():
            if ledger.get(k) != v:
                fail(f"ledger mismatch for {k!r}: expected {v} got {ledger.get(k)}")
    else:
        warn("source file differs from the known PHIVOLCS workbook; ledger reported but not asserted")

    # ── 7. Spot checks ──
    rng = random.Random(42)
    sample = [events[0], events[-1]] + rng.sample(events, 5)
    spot = []
    for ev in sample:
        spot.append({
            "id": ev["id"], "sheet": ev["source_sheet"], "row": ev["source_row"],
            "datetime_utc": ev["datetime_utc"], "lat": ev["latitude"], "lon": ev["longitude"],
            "depth_km": ev["depth_km"], "mag": ev["mag"], "mag_type": ev["mag_type"],
            "qa_flags": ev["qa_flags"],
        })

    status = "pass" if not failures else "fail"
    REPORTS.mkdir(parents=True, exist_ok=True)
    audit_json = {
        "status": status,
        "generated_utc": _dt.datetime.now(_dt.timezone.utc).isoformat(),
        "source_file": XLSX.name,
        "source_sha256": actual_sha,
        "counts": {
            "events": len(events),
            "geojson_features": len(feats),
            "xlsx_data_rows": n_data_rows_total,
            "nonempty_cells_seen": n_cells_nonempty,
            "data_cells_reconciled": n_cells_reconciled,
            "max_row_width_seen": ghost_max_col_seen,
        },
        "ledger": ledger,
        "failures": failures,
        "warnings": warnings,
        "info": info,
        "spot_checks": spot,
    }
    (REPORTS / "audit.json").write_text(
        json.dumps(audit_json, indent=2, ensure_ascii=False), encoding="utf-8")

    lines = [
        "# PHIVOLCS Catalogue Capture Audit",
        "",
        f"**Status: {'PASS' if status == 'pass' else 'FAIL'}**",
        "",
        f"- Source: `{XLSX.name}` (sha256 `{actual_sha[:16]}...`)",
        f"- Generated: {audit_json['generated_utc']}",
        "",
        "## Counts",
        f"- Events in catalog.json: **{len(events)}**",
        f"- Features in catalog.geojson: **{len(feats)}**",
        f"- Data rows found in xlsx: **{n_data_rows_total}**",
        f"- Per sheet: " + ", ".join(f"`{k}` = {v}" for k, v in ledger["per_sheet"].items()),
        f"- Non-empty cells seen in workbook: {n_cells_nonempty} "
        f"(widest row scanned: {ghost_max_col_seen} columns)",
        f"- Data cells reconciled 1:1 against events: {n_cells_reconciled}",
        "",
        "## Partition proof",
        "Every non-empty workbook cell was matched to exactly one of: preamble metadata, "
        "header metadata, or an event field. Unclaimed cells: "
        f"**{len(unclaimed)}**; metadata cells missing from workbook: **{len(claimed_missing)}**; "
        f"cell/field mismatches: **{len(recon_mismatches)}**.",
        "",
        "## Anomaly ledger" + (" (asserted against known source)" if known else " (report only)"),
        f"- QA flags: {json.dumps(flag_counts)}",
        f"- Invalid datetimes: {ledger['invalid_datetime_rows']} (Hour=27 kept raw, datetime null)",
        f"- Preferred magnitude distribution: {json.dumps(mag_type_counts)}",
        f"- Events with preferred mag < 5.0: {sub_m5}",
        f"- Exact duplicate time+location pairs: {len(dups)}",
        "",
        "## Ranges",
        f"- Latitude: {min(lats)} to {max(lats)} degN",
        f"- Longitude: {min(lons)} to {max(lons)} degE",
        f"- Depth: {min(depths)} to {max(depths)} km",
        f"- Preferred magnitude: {min(mags)} to {max(mags)}",
        f"- Years: {meta['year_min']} to {meta['year_max']}",
        "",
        "## Spot checks (first, last, 5 seeded-random)",
    ]
    for s in spot:
        lines.append(f"- `{s['id']}` {s['sheet']} row {s['row']}: {s['datetime_utc']} "
                     f"lat {s['lat']} lon {s['lon']} depth {s['depth_km']} km "
                     f"M{s['mag']} ({s['mag_type']}) flags={s['qa_flags']}")
    if warnings:
        lines += ["", "## Warnings"] + [f"- {w}" for w in warnings]
    if failures:
        lines += ["", "## FAILURES"] + [f"- {f}" for f in failures]
    if info:
        lines += ["", "## Notes"] + [f"- {i}" for i in info]
    (REPORTS / "audit_report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(f"AUDIT {status.upper()}: {len(events)} events, {n_cells_reconciled} cells reconciled, "
          f"{len(unclaimed)} unclaimed, {len(recon_mismatches)} mismatches, "
          f"{len(failures)} failures, {len(warnings)} warnings")
    for f_ in failures[:20]:
        print("  FAIL:", f_)
    for w in warnings:
        print("  warn:", w)
    sys.exit(0 if status == "pass" else 1)


if __name__ == "__main__":
    main()
